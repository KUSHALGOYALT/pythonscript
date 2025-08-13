#!/usr/bin/env python3
import os
import logging
import requests
import re
from datetime import datetime, timedelta
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

NRPC_BASE_URL = "http://164.100.60.165"
NRPC_DSA_URL = "http://164.100.60.165/comm/dsa.html"
NRLDC_DIR = "../data"

def setup_directories():
    os.makedirs(NRLDC_DIR, exist_ok=True)
    logging.info(f"✅ Created directory: {NRLDC_DIR}")

def try_download_historical_files():
    logging.info("🔍 Trying to download historical files for past 7 days...")
    
    downloaded_files = []
    current_date = datetime.now()
    
    for days_back in range(7):
        test_date = current_date - timedelta(days=days_back)
        date_str = test_date.strftime('%d%m%Y')
        
        file_patterns = [
            f"Supporting_files_{date_str}.xls",
            f"Supporting_files_{date_str}.xlsx",
            f"DSA_{date_str}.xls",
            f"DSA_{date_str}.xlsx",
            f"NRLDC_DSM_{date_str}.xls",
            f"NRLDC_DSM_{date_str}.xlsx",
            f"Supporting_Files_{date_str}.xls",
            f"Supporting_Files_{date_str}.xlsx"
        ]
        
        for filename in file_patterns:
            possible_urls = [
                f"{NRPC_BASE_URL}/comm/2021-22/dsa/{date_str}/Supporting_files.xls",
                f"{NRPC_BASE_URL}/comm/2022-23/dsa/{date_str}/Supporting_files.xls",
                f"{NRPC_BASE_URL}/comm/2023-24/dsa/{date_str}/Supporting_files.xls",
                f"{NRPC_BASE_URL}/comm/2024-25/dsa/{date_str}/Supporting_files.xls",
                f"{NRPC_BASE_URL}/comm/dsa/{date_str}/Supporting_files.xls",
                f"{NRPC_BASE_URL}/files/{filename}",
                f"{NRPC_BASE_URL}/data/{filename}"
            ]
            
            for url in possible_urls:
                try:
                    logging.debug(f"🔍 Testing: {url}")
                    response = requests.get(url, timeout=10, verify=True)
                    
                    if response.status_code == 200 and len(response.content) > 1000:
                        file_path = os.path.join(NRLDC_DIR, filename)
                        with open(file_path, 'wb') as f:
                            f.write(response.content)
                        
                        logging.info(f"✅ Successfully downloaded: {filename} ({len(response.content)} bytes)")
                        downloaded_files.append({
                            'url': url,
                            'filename': filename,
                            'type': 'excel'
                        })
                        break
                        
                except Exception as e:
                    logging.debug(f"⚠️ Failed to access {url}: {e}")
                    continue
            
            if any(f"{date_str}" in f['filename'] for f in downloaded_files):
                break
    
    return downloaded_files

def get_dsa_links():
    """Get all DSA links from the dsa.html page"""
    try:
        logging.info("🔍 Accessing NRPC DSA page...")
        logging.info(f"🔍 Accessing: {NRPC_DSA_URL}")
        
        response = requests.get(NRPC_DSA_URL, timeout=30)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Find all DSA links from the dropdown options
        dsa_links = []
        
        # Look for select options that contain DSA week patterns
        select_element = soup.find('select', {'name': 'wk'})
        if select_element:
            for option in select_element.find_all('option'):
                value = option.get('value', '')
                text = option.get_text(strip=True)
                
                # Look for DSA week patterns (like 210725-270725(WK-17))
                if re.search(r'\d{6}-\d{6}\(WK-\d+\)', value):
                    # Construct the URL based on the JavaScript pattern we saw
                    dsa_url = f"comm/2021-22/dsa/{value}/Supporting_files.xls"
                    full_url = urljoin(NRPC_BASE_URL, dsa_url)
                    
                    dsa_links.append({
                        'text': text,
                        'value': value,
                        'url': full_url
                    })
                    logging.info(f"Found DSA link: {text} -> {value}")
        
        return dsa_links
        
    except Exception as e:
        logging.error(f"❌ Error getting DSA links: {e}")
        return []

def find_supporting_files(dsa_url):
    """Find Supporting_files.xls in a DSA directory"""
    try:
        # Extract the directory path from the DSA URL
        if '/dsa/' in dsa_url:
            base_dir = dsa_url.rsplit('/', 1)[0]  # Remove the filename
            supporting_urls = [
                f"{base_dir}/Supporting_files.xls",
                f"{base_dir}/Supporting_files.xlsx",
                f"{base_dir}/supporting_files.xls",
                f"{base_dir}/supporting_files.xlsx",
                f"{base_dir}/Supporting_Files.xls",
                f"{base_dir}/Supporting_Files.xlsx"
            ]
            
            for supporting_url in supporting_urls:
                try:
                    logging.info(f"🔍 Checking: {supporting_url}")
                    response = requests.head(supporting_url, timeout=10)
                    if response.status_code == 200:
                        logging.info(f"✅ Found supporting file: {supporting_url}")
                        return supporting_url
                except Exception as e:
                    logging.debug(f"⚠️ Error checking {supporting_url}: {e}")
                    continue
        
        return None
        
    except Exception as e:
        logging.error(f"❌ Error finding supporting files: {e}")
        return None

def download_file(url, filename):
    """Download a file from URL"""
    try:
        logging.info(f"📥 Downloading: {filename}")
        
        response = requests.get(url, timeout=60, stream=True)
        response.raise_for_status()
        
        # Extract just the filename from the URL if a full URL is passed
        if filename.startswith('http'):
            from urllib.parse import urlparse
            filename = os.path.basename(urlparse(filename).path)
        
        file_path = os.path.join(NRLDC_DIR, filename)
        
        with open(file_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
        
        logging.info(f"✅ Downloaded: {filename}")
        return file_path
        
    except Exception as e:
        logging.error(f"❌ Error downloading {filename}: {e}")
        return None

def process_excel_file(file_path):
    """Process Excel file using openpyxl or xlrd"""
    try:
        logging.info(f"Processing file: {file_path}")
        
        file_extension = os.path.splitext(file_path)[1].lower()
        
        if file_extension == '.xlsx':
            # Use openpyxl for .xlsx files
            workbook = load_workbook(file_path, data_only=True)
            sheet_names = workbook.sheetnames
            logging.info(f"Found {len(sheet_names)} sheets: {sheet_names}")
            
            # Process each sheet
            processed_data = {}
            for sheet_name in sheet_names:
                try:
                    logging.info(f"Reading sheet: {sheet_name}")
                    worksheet = workbook[sheet_name]
                    
                    # Convert worksheet to list of lists
                    data = []
                    for row in worksheet.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):  # Skip completely empty rows
                            data.append(row)
                    
                    if data:
                        # Find the last non-empty column
                        max_col = 0
                        for row in data:
                            for i, cell in enumerate(row):
                                if cell is not None:
                                    max_col = max(max_col, i)
                        
                        # Clean data - remove empty rows and columns
                        cleaned_data = []
                        for row in data:
                            if any(cell is not None for cell in row[:max_col+1]):
                                cleaned_row = [str(cell).strip() if cell is not None else '' for cell in row[:max_col+1]]
                                cleaned_data.append(cleaned_row)
                        
                        if cleaned_data:
                            processed_data[sheet_name] = cleaned_data
                            logging.info(f"✅ Processed sheet '{sheet_name}' with {len(cleaned_data)} rows and {len(cleaned_data[0]) if cleaned_data else 0} columns")
                        else:
                            logging.warning(f"Sheet '{sheet_name}' is empty after cleaning")
                    else:
                        logging.warning(f"Sheet '{sheet_name}' is empty")
                        
                except Exception as e:
                    logging.error(f"❌ Error processing sheet '{sheet_name}': {e}")
                    continue
            
            return processed_data
            
        elif file_extension == '.xls':
            # Use xlrd for .xls files
            import xlrd
            workbook = xlrd.open_workbook(file_path)
            sheet_names = workbook.sheet_names()
            logging.info(f"Found {len(sheet_names)} sheets: {sheet_names}")
            
            # Process each sheet
            processed_data = {}
            for sheet_name in sheet_names:
                try:
                    logging.info(f"Reading sheet: {sheet_name}")
                    worksheet = workbook.sheet_by_name(sheet_name)
                    
                    # Convert worksheet to list of lists
                    data = []
                    for row_idx in range(worksheet.nrows):
                        row_data = []
                        for col_idx in range(worksheet.ncols):
                            cell_value = worksheet.cell_value(row_idx, col_idx)
                            row_data.append(cell_value)
                        
                        if any(cell_value != '' for cell_value in row_data):  # Skip completely empty rows
                            data.append(row_data)
                    
                    if data:
                        # Find the last non-empty column
                        max_col = 0
                        for row in data:
                            for i, cell in enumerate(row):
                                if cell != '':
                                    max_col = max(max_col, i)
                        
                        # Clean data - remove empty rows and columns
                        cleaned_data = []
                        for row in data:
                            if any(cell != '' for cell in row[:max_col+1]):
                                cleaned_row = [str(cell).strip() if cell != '' else '' for cell in row[:max_col+1]]
                                cleaned_data.append(cleaned_row)
                        
                        if cleaned_data:
                            processed_data[sheet_name] = cleaned_data
                            logging.info(f"✅ Processed sheet '{sheet_name}' with {len(cleaned_data)} rows and {len(cleaned_data[0]) if cleaned_data else 0} columns")
                        else:
                            logging.warning(f"Sheet '{sheet_name}' is empty after cleaning")
                    else:
                        logging.warning(f"Sheet '{sheet_name}' is empty")
                        
                except Exception as e:
                    logging.error(f"❌ Error processing sheet '{sheet_name}': {e}")
                    continue
            
            return processed_data
        
        else:
            logging.error(f"❌ Unsupported file format: {file_extension}")
            return None
        
    except Exception as e:
        logging.error(f"❌ Error processing file: {e}")
        return None

def save_processed_data(processed_data, filename):
    """Save processed data to Excel file using openpyxl"""
    if not processed_data:
        logging.warning("No data to save")
        return None
    
    try:
        # Generate output filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_name = os.path.splitext(filename)[0]
        output_filename = f"{base_name}_processed_{timestamp}.xlsx"
        output_path = os.path.join(NRLDC_DIR, output_filename)
        
        # Create new workbook
        workbook = Workbook()
        # Remove default sheet
        workbook.remove(workbook.active)
        
        for sheet_name, data in processed_data.items():
            # Clean sheet name for Excel (remove invalid characters)
            clean_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)
            
            # Create new sheet
            worksheet = workbook.create_sheet(title=clean_sheet_name)
            
            # Write data to sheet
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
            
            logging.info(f"✅ Added sheet '{clean_sheet_name}' with {len(data)} rows")
        
        # Save workbook
        workbook.save(output_path)
        
        logging.info(f"✅ Saved processed data to: {output_path}")
        return output_path
        
    except Exception as e:
        logging.error(f"❌ Error saving processed data: {e}")
        return None

def main():
    """Main function"""
    try:
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        
        logging.info("🚀 Starting NRLDC DSA data extraction...")
        
        # Setup directories
        setup_directories()
        
        # First, try to download historical files for past 7 days
        historical_files = try_download_historical_files()
        if historical_files:
            logging.info(f"📥 Downloaded {len(historical_files)} historical files")
            for file_info in historical_files:
                try:
                    file_path = os.path.join(NRLDC_DIR, file_info['filename'])
                    if os.path.exists(file_path):
                        # Process the Excel file
                        processed_data = process_excel_file(file_path)
                        
                        if processed_data:
                            # Save processed data
                            output_path = save_processed_data(processed_data, file_info['filename'])
                            if output_path:
                                logging.info(f"✅ Processed historical file: {file_info['filename']} -> {output_path}")
                except Exception as e:
                    logging.error(f"❌ Error processing historical file {file_info['filename']}: {e}")
                    continue
        
        # Get DSA links from the dsa.html page
        dsa_links = get_dsa_links()
        
        if not dsa_links:
            logging.error("❌ No DSA links found")
            return
        
        logging.info(f"✅ Found {len(dsa_links)} DSA links")
        
        # Process each DSA link to download supporting files
        downloaded_files = []
        
        for i, dsa_link in enumerate(dsa_links[:5]):  # Limit to first 5 to avoid overwhelming
            try:
                logging.info(f"🔍 Processing DSA {i+1}/{min(5, len(dsa_links))}: {dsa_link['text']}")
                
                # Download the supporting file directly
                filename = f"Supporting_files_{dsa_link['value']}.xls"
                file_path = download_file(dsa_link['url'], filename)
                
                if file_path:
                    # Process the Excel file
                    processed_data = process_excel_file(file_path)
                    
                    if processed_data:
                        # Save processed data
                        output_path = save_processed_data(processed_data, filename)
                        if output_path:
                            downloaded_files.append({
                                'dsa': dsa_link['text'],
                                'file': filename,
                                'output': output_path
                            })
                
            except Exception as e:
                logging.error(f"❌ Error processing DSA {dsa_link['text']}: {e}")
                continue
        
        if downloaded_files:
            logging.info(f"🎉 DSA extraction completed successfully!")
            logging.info(f"📊 Downloaded and processed {len(downloaded_files)} files:")
            for file_info in downloaded_files:
                logging.info(f"   • {file_info['dsa']} -> {file_info['output']}")
        else:
            logging.warning("⚠️ No supporting files were successfully downloaded")
        
    except Exception as e:
        logging.error(f"❌ Error in main function: {e}")

if __name__ == "__main__":
    main()
