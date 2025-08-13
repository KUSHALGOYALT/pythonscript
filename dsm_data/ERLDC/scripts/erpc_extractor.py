#!/usr/bin/env python3
import os
import sys
import logging
import requests
from datetime import datetime, timedelta
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('erpc_extractor.log'),
        logging.StreamHandler()
    ]
)

ERPC_BASE_URL = "https://erpc.gov.in"
ERPC_DIR = "../data"

def setup_directories():
    os.makedirs(ERPC_DIR, exist_ok=True)
    logging.info(f"✅ Created directory: {ERPC_DIR}")

def try_download_historical_files():
    logging.info("🔍 Trying to download historical files for past 7 days...")
    
    downloaded_files = []
    current_date = datetime.now()
    
    for days_back in range(7):
        test_date = current_date - timedelta(days=days_back)
        date_str = test_date.strftime('%d%m%Y')
        
        file_patterns = [
            f"ERPC_DSM_{date_str}.xlsx",
            f"ERPC_DSM_{date_str}.xls",
            f"ERPC_UI_{date_str}.xlsx",
            f"ERPC_UI_{date_str}.xls",
            f"ERPC_Deviation_{date_str}.xlsx",
            f"ERPC_Deviation_{date_str}.xls",
            f"DSM_{date_str}.xlsx",
            f"DSM_{date_str}.xls",
            f"UI_{date_str}.xlsx",
            f"UI_{date_str}.xls"
        ]
        
        for filename in file_patterns:
            possible_urls = [
                f"{ERPC_BASE_URL}/uploads/{filename}",
                f"{ERPC_BASE_URL}/en/ui-and-deviation-accts/{filename}",
                f"{ERPC_BASE_URL}/files/{filename}",
                f"{ERPC_BASE_URL}/data/{filename}"
            ]
            
            for url in possible_urls:
                try:
                    logging.debug(f"🔍 Testing: {url}")
                    response = requests.get(url, timeout=10, verify=True)
                    
                    if response.status_code == 200 and len(response.content) > 1000:
                        file_path = os.path.join(ERPC_DIR, filename)
                        with open(file_path, 'wb') as f:
                            f.write(response.content)
                        
                        logging.info(f"✅ Successfully downloaded: {filename} ({len(response.content)} bytes)")
                        downloaded_files.append({
                            'url': url,
                            'filename': filename,
                            'type': 'excel' if filename.endswith(('.xlsx', '.xls')) else 'csv'
                        })
                        break
                        
                except Exception as e:
                    logging.debug(f"⚠️ Failed to access {url}: {e}")
                    continue
            
            if any(f"{date_str}" in f['filename'] for f in downloaded_files):
                break
    
    return downloaded_files

def scrape_erpc_website():
    """Scrape the ERPC website for data files"""
    logging.info("🌐 Scraping ERPC website...")
    
    try:
        # Get the main page
        response = requests.get(ERPC_BASE_URL, timeout=30)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Look for uploads directory
        uploads_links = []
        for link in soup.find_all('a', href=True):
            href = link['href']
            if 'uploads' in href and any(ext in href.lower() for ext in ['.xlsx', '.xls', '.csv']):
                full_url = urljoin(ERPC_BASE_URL, href)
                uploads_links.append({
                    'url': full_url,
                    'href': href,
                    'text': link.get_text(strip=True),
                    'type': 'excel' if href.endswith(('.xlsx', '.xls')) else 'csv'
                })
                logging.info(f"🎯 Found Excel file in uploads: {link.get_text(strip=True)} -> {full_url}")
        
        # Categorize files
        categorized_files = {
            'excel': uploads_links,
            'csv': []
        }
        
        # Also check subdirectories
        subdirs = ['/en/ui-and-deviation-accts/', '/uploads/']
        for subdir in subdirs:
            try:
                sub_url = urljoin(ERPC_BASE_URL, subdir)
                logging.info(f"🔍 Found potential subdirectory: {link.get_text(strip=True)} -> {sub_url}")
                
                sub_resp = requests.get(sub_url, timeout=30, verify=True)
                if sub_resp.status_code == 200:
                    sub_soup = BeautifulSoup(sub_resp.content, 'html.parser')
                    
                    for link in sub_soup.find_all('a', href=True):
                        href = link['href']
                        if any(ext in href.lower() for ext in ['.xlsx', '.xls', '.csv']):
                            full_url = urljoin(sub_url, href)
                            categorized_files['excel'].append({
                                'url': full_url,
                                'href': href,
                                'text': link.get_text(strip=True),
                                'type': 'excel' if href.endswith(('.xlsx', '.xls')) else 'csv'
                            })
                            logging.info(f"🎯 Found Excel file in subdirectory: {link.get_text(strip=True)} -> {full_url}")
                            
            except Exception as e:
                logging.debug(f"⚠️ Error checking subdirectory {subdir}: {e}")
                continue
        
        return categorized_files
        
    except Exception as e:
        logging.error(f"❌ Error scraping ERPC website: {e}")
        return {'excel': [], 'csv': []}

def download_file(url, filename):
    """Download a file from URL"""
    try:
        logging.info(f"📥 Downloading: {filename}")
        
        response = requests.get(url, timeout=30, stream=True)
        response.raise_for_status()
        
        # Extract just the filename from the URL if a full URL is passed
        if filename.startswith('http'):
            from urllib.parse import urlparse
            filename = os.path.basename(urlparse(filename).path)
        
        file_path = os.path.join(ERPC_DIR, filename)
        
        with open(file_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
        
        logging.info(f"✅ Downloaded: {filename}")
        return file_path
        
    except Exception as e:
        logging.error(f"❌ Error downloading {filename}: {e}")
        return None

def process_excel_file(file_path, filename):
    """Process Excel file using openpyxl"""
    try:
        logging.info(f"📊 Processing Excel file: {filename}")
        
        # Load workbook
        workbook = load_workbook(file_path, data_only=True)
        sheet_names = workbook.sheetnames
        logging.info(f"Found {len(sheet_names)} sheets: {sheet_names}")
        
        processed_data = {}
        for sheet_name in sheet_names:
            try:
                logging.info(f"Reading sheet: {sheet_name}")
                worksheet = workbook[sheet_name]
                
                # Convert worksheet to list of lists
                data = []
                for row in worksheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):  # Skip completely empty rows
                        data.append(row)
                
                if data:
                    # Find the last non-empty column
                    max_col = 0
                    for row in data:
                        for i, cell in enumerate(row):
                            if cell is not None:
                                max_col = max(max_col, i)
                    
                    # Clean data - remove empty rows and columns
                    cleaned_data = []
                    for row in data:
                        if any(cell is not None for cell in row[:max_col+1]):
                            cleaned_row = [str(cell).strip() if cell is not None else '' for cell in row[:max_col+1]]
                            cleaned_data.append(cleaned_row)
                    
                    if cleaned_data:
                        processed_data[sheet_name] = cleaned_data
                        logging.info(f"✅ Processed sheet '{sheet_name}' with {len(cleaned_data)} rows")
                    else:
                        logging.warning(f"⚠️ Sheet '{sheet_name}' is empty after cleaning")
                else:
                    logging.warning(f"⚠️ Sheet '{sheet_name}' is empty")
                    
            except Exception as e:
                logging.error(f"❌ Error processing sheet '{sheet_name}': {e}")
                continue
        
        return processed_data
        
    except Exception as e:
        logging.error(f"❌ Error processing Excel file {file_path}: {e}")
        return None

def save_processed_data(processed_data, original_filename):
    """Save processed data to Excel file using openpyxl"""
    if not processed_data:
        logging.warning("No data to save")
        return None
    
    try:
        target_dir = ERPC_DIR
        logging.info(f"📁 Saving ERPC data to: {target_dir}")
        
        base_name = os.path.splitext(original_filename)[0]
        output_filename = f"{base_name}_processed.xlsx"
        output_path = os.path.join(target_dir, output_filename)
        
        # Create new workbook
        workbook = Workbook()
        # Remove default sheet
        workbook.remove(workbook.active)
        
        for sheet_name, data in processed_data.items():
            # Create new sheet
            worksheet = workbook.create_sheet(title=sheet_name[:31])  # Excel sheet name limit
            
            # Write data to sheet
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
            
            logging.info(f"💾 Saved sheet '{sheet_name}' with {len(data)} rows")
        
        # Save workbook
        workbook.save(output_path)
        logging.info(f"✅ Saved processed data to: {output_path}")
        return output_path
        
    except Exception as e:
        logging.error(f"❌ Error saving processed data: {e}")
        return None

def download_and_process_file(file_info, data_type):
    # Extract proper filename from URL
    from urllib.parse import urlparse
    url = file_info['url']
    filename = os.path.basename(urlparse(url).path)
    
    file_path = download_file(url, filename)
    if not file_path:
        return False
    
    if file_info['type'] == 'excel':
        processed_data = process_excel_file(file_path, filename)
        if processed_data:
            output_path = save_processed_data(processed_data, filename)
            if output_path:
                return True
    
    return False

def run_update():
    logging.info("🔄 Starting ERPC data update...")
    
    try:
        total_processed = 0
        
        # First, try to download historical files for past 7 days
        historical_files = try_download_historical_files()
        if historical_files:
            logging.info(f"📥 Downloaded {len(historical_files)} historical files")
            for file_info in historical_files:
                if download_and_process_file(file_info, file_info['type']):
                    total_processed += 1
        
        # Then scrape current website for additional files
        categorized_files = scrape_erpc_website()
        
        for data_type, files in categorized_files.items():
            if files:
                logging.info(f"Processing {data_type} files...")
                for file_info in files:
                    if download_and_process_file(file_info, data_type):
                        total_processed += 1
        
        logging.info(f"✅ Update completed. Processed {total_processed} files.")
        
    except Exception as e:
        logging.error(f"❌ Error in update: {e}")

def main():
    if len(sys.argv) > 1:
        if sys.argv[1] == '--help':
            print("ERPC Data Extractor Usage:")
            print("  python erpc_extractor.py          # Run data extraction")
            print("  python erpc_extractor.py --help   # Show this help")
            return
        else:
            print("Unknown argument. Use --help for usage information.")
            return
    
    setup_directories()
    run_update()

if __name__ == "__main__":
    main()
